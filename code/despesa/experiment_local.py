"""
Experimento de aprendizado local: agrupa as séries por similaridade e treina
modelos no pool de cada cluster (Ridge, SVR e MLP pooled, mais um LightGBM por
cluster).

Ordem para não vazar (o conjunto de teste não entra nos passos de features e treino):
divide treino/conjunto de teste com os mesmos splits do experiment.py, extrai as features
de similaridade só do treino e clusteriza por silhouette. Para cada série
não-ruído-branco, monta o pool com os treinos do cluster, treina os modelos (com
CV interno no pool) e avalia no conjunto de teste de forma recursiva. No fim, carrega os
resultados individuais do experiment.py, compara individual vs local e exporta
CSV + Excel.

Rodar com: python experiment_local.py
"""

from __future__ import annotations

import glob
import logging
import pickle
import sys
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import wilcoxon

warnings.filterwarnings("ignore")

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

from config import (
    CAMINHO_PARQUET, RESULTS_DIR,
    LR_ALPHA, RANDOM_STATE, MESES_PREV,
)
from utils import (
    carregar_dados, construir_series, mapear_nomes,
    filtrar_series, preparar_serie,
    dividir_serie,
)
from utils.acf_analysis import analisar_serie
from utils.feature_extraction import extrair_features_todas
from clustering import clusterizar_series
from models.local_model import (
    treinar_avaliar_pooled_ridge,
    treinar_avaliar_pooled_svr,
    treinar_avaliar_pooled_mlp,
)
from models.lgbm_cluster import treinar_avaliar_lgbm_cluster
from reports.plots import plotar_serie

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s -- %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("experiment_local")

# Checkpoint
_CHECKPOINT_VERSION = "v1_liquidado_local_multistep_v2"  # v2: grids unificados + busca de alpha Ridge
_CHECKPOINT_PATH    = RESULTS_DIR / "checkpoint_local.pkl"

def _salvar_checkpoint_local(registros, timestamp, concluidos):
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    data = {
        "registros"      : registros,
        "timestamp"      : timestamp,
        "concluidos"     : concluidos,
        "schema_version" : _CHECKPOINT_VERSION,
    }
    with open(_CHECKPOINT_PATH, "wb") as f:
        pickle.dump(data, f)

def _carregar_checkpoint_local():
    if not _CHECKPOINT_PATH.exists():
        return None
    try:
        with open(_CHECKPOINT_PATH, "rb") as f:
            data = pickle.load(f)
        if data.get("schema_version") != _CHECKPOINT_VERSION:
            logger.warning("Checkpoint local com schema incompatível. Iniciando do zero.")
            return None
        logger.info("Checkpoint local carregado: %d séries já concluídas.", len(data["concluidos"]))
        return data
    except Exception as exc:
        logger.warning("Falha ao carregar checkpoint local: %s. Iniciando do zero.", exc)
        return None


# Helpers

def _fmt(v: float) -> str:
    if pd.isna(v):    return "N/A"
    if abs(v) >= 1e9: return f"R$ {v/1e9:.2f} Bi"
    if abs(v) >= 1e6: return f"R$ {v/1e6:.2f} Mi"
    if abs(v) >= 1e3: return f"R$ {v/1e3:.2f} K"
    return f"R$ {v:,.2f}"


def _carregar_resultados_individuais() -> pd.DataFrame | None:
    """Carrega o CSV de metricas mais recente do experiment.py."""
    csvs = sorted(glob.glob(str(RESULTS_DIR / "metricas_*.csv")))
    if not csvs:
        logger.warning("Nenhum CSV de metricas encontrado em results/. "
                       "Execute experiment.py primeiro.")
        return None
    path = csvs[-1]
    logger.info("Carregando resultados individuais de: %s", Path(path).name)
    return pd.read_csv(path)


def _wilcoxon_safe(a: np.ndarray, b: np.ndarray) -> tuple[float, float]:
    """Wilcoxon signed-rank; retorna (stat, p) -- NaN se n < 5."""
    diff = a - b
    diff = diff[~np.isnan(diff)]
    if len(diff) < 5:
        return float("nan"), float("nan")
    try:
        stat, p = wilcoxon(diff, alternative="two-sided", zero_method="wilcox")
        return float(stat), float(p)
    except Exception:
        return float("nan"), float("nan")


# Orquestrador principal

def rodar_experimento_local(
    caminho_parquet: Path = CAMINHO_PARQUET,
    k_min: int = 2,
    k_max: int = 6,
    verbose: bool = True,
    gerar_graficos: bool = True,
) -> dict:
    """
    Executa o experimento de aprendizado local completo.

    Returns
    -------
    dict com:
      comparacao_df   -- DataFrame de comparacao por serie
      clustering      -- ResultadoClustering
      features_df     -- features de similaridade
      timestamp       -- string do timestamp
    """
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # Tenta retomar checkpoint
    ckpt = _carregar_checkpoint_local()
    if ckpt:
        timestamp  = ckpt["timestamp"]
        registros  = ckpt["registros"]
        concluidos = ckpt["concluidos"]
    else:
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        registros  = []
        concluidos = set()

    # 1. Carrega e divide series
    logger.info("Carregando dados de %s", caminho_parquet)
    df         = carregar_dados(caminho_parquet)
    series_raw = construir_series(df)
    nomes      = mapear_nomes(df)
    series, _  = filtrar_series(series_raw)

    logger.info("Series validas: %d", len(series))

    # Divide TODAS as series em treino/conjunto de teste (necessario para features e pool)
    treinos         : dict = {}
    testes          : dict = {}
    treino_indices  : dict = {}
    teste_indices   : dict = {}
    acf_infos       : dict = {}
    nlags_map       : dict = {}
    grupos          : dict = {}
    shifts          : dict = {}

    for cod, serie_original in series.items():
        serie_pronta, grupo_sinal, shift_value = preparar_serie(serie_original)
        grupos[cod] = grupo_sinal
        shifts[cod] = shift_value

        (treino_vals, teste_vals,
         treino_idx, teste_idx,
         n_holdout, n_lags) = dividir_serie(serie_pronta)

        treinos[cod]        = treino_vals
        testes[cod]         = teste_vals
        treino_indices[cod] = treino_idx
        teste_indices[cod]  = teste_idx

        # ACF/PACF no treino -- define n_lags e ruido branco
        acf_info        = analisar_serie(treino_vals)
        acf_infos[cod]  = acf_info
        nlags_map[cod]  = acf_info["n_lags"] if not acf_info["e_ruido_branco"] else n_holdout

    # Series modelaveis (nao-ruido-branco)
    series_modelaveis = [
        cod for cod, info in acf_infos.items()
        if not info["e_ruido_branco"]
    ]
    series_modelaveis_set = set(series_modelaveis)   # lookup O(1) no loop
    logger.info("Series modelaveis (nao-RB): %d / %d",
                len(series_modelaveis), len(series))

    # 2. Extrai features de similaridade (somente do treino)
    logger.info("Extraindo features de similaridade (somente do treino)...")
    features_df = extrair_features_todas(treinos)

    # 3. Clustering
    logger.info("Clusterizando %d series (k em [%d, %d])...",
                len(features_df), k_min, k_max)
    resultado_clustering = clusterizar_series(
        features_df, k_min=k_min, k_max=k_max, random_state=RANDOM_STATE
    )

    if verbose:
        print(f"\nClustering: k={resultado_clustering.k}  "
              f"silhouette={resultado_clustering.silhouette:.4f}")
        for cid, membros in resultado_clustering.cluster_members.items():
            n_mod = sum(1 for m in membros if m in series_modelaveis)
            print(f"  Cluster {cid}: {len(membros)} series "
                  f"({n_mod} modelaveis)")

    # 4. Treina e avalia modelos pooled por serie modelavel
    total = len(series_modelaveis)

    for idx, cod in enumerate(series_modelaveis, 1):
        # Pula séries já concluídas (checkpoint)
        if str(cod) in concluidos:
            if verbose:
                print(f"[{idx}/{total}] Fonte {cod} — já concluída, pulando.")
            continue

        nome           = nomes.get(cod, "---")
        cluster_id     = resultado_clustering.cluster_map[cod]
        membros_cl     = resultado_clustering.cluster_members[cluster_id]
        # Filtra RB do pool — apenas séries modeláveis contribuem ao treino
        membros_cl_mod = [c for c in membros_cl if c in series_modelaveis_set]
        n_lags         = nlags_map[cod]
        teste_vals     = testes[cod]

        if verbose:
            print(f"\n[{idx}/{total}] Fonte {cod} | cluster={cluster_id} "
                  f"({len(membros_cl_mod)}/{len(membros_cl)} modelaveis) | n_lags={n_lags}")
            print(f"  {nome[:70]}")

        _nan = {"RMSE": float("nan"), "MAE": float("nan"),
                "MAPE": float("nan"), "params": "erro", "preds": []}

        # Ridge pooled
        logger.info("[%s] Pooled-Ridge ...", cod)
        try:
            r_ridge = treinar_avaliar_pooled_ridge(
                membros_cl_mod, treinos, cod, teste_vals, n_lags
            )  # alpha=None: busca CV via GRID_RIDGE_ALPHAS
        except Exception as exc:
            logger.error("[%s] Ridge falhou: %s", cod, exc)
            r_ridge = dict(_nan)
        if verbose:
            print(f"  Ridge  RMSE={r_ridge['RMSE']:>14,.2f}  "
                  f"MAPE={r_ridge['MAPE']:.2f}%")

        # SVR pooled
        logger.info("[%s] Pooled-SVR ...", cod)
        try:
            r_svr = treinar_avaliar_pooled_svr(
                membros_cl_mod, treinos, cod, teste_vals, n_lags,
                random_state=RANDOM_STATE,
            )
        except Exception as exc:
            logger.error("[%s] SVR falhou: %s", cod, exc)
            r_svr = dict(_nan)
        if verbose:
            print(f"  SVR    RMSE={r_svr['RMSE']:>14,.2f}  "
                  f"MAPE={r_svr['MAPE']:.2f}%")

        # MLP pooled
        logger.info("[%s] Pooled-MLP ...", cod)
        try:
            r_mlp = treinar_avaliar_pooled_mlp(
                membros_cl_mod, treinos, cod, teste_vals, n_lags,
                random_state=RANDOM_STATE,
            )
        except Exception as exc:
            logger.error("[%s] MLP falhou: %s", cod, exc)
            r_mlp = dict(_nan)
        if verbose:
            print(f"  MLP    RMSE={r_mlp['RMSE']:>14,.2f}  "
                  f"MAPE={r_mlp['MAPE']:.2f}%")

        # LGBM cluster
        logger.info("[%s] LGBM-Cluster ...", cod)
        try:
            r_lgbm_cl = treinar_avaliar_lgbm_cluster(
                membros_cl_mod   = membros_cl_mod,
                treinos          = treinos,
                treino_indices   = treino_indices,
                codigo_alvo      = cod,
                teste_vals       = teste_vals,
                teste_idx        = teste_indices[cod],
                random_state     = RANDOM_STATE,
            )
        except Exception as exc:
            logger.error("[%s] LGBM-Cluster falhou: %s", cod, exc)
            r_lgbm_cl = dict(_nan)
        if verbose:
            rmse_str = f"{r_lgbm_cl['RMSE']:>14,.2f}" if not np.isnan(r_lgbm_cl['RMSE']) else "           N/A"
            mape_str = f"{r_lgbm_cl['MAPE']:.2f}%"   if not np.isnan(r_lgbm_cl['MAPE']) else "N/A"
            print(f"  LGBM_Cl RMSE={rmse_str}  MAPE={mape_str}")

        acf_info = acf_infos[cod]

        # Checkpoint após cada série
        concluidos.add(str(cod))

        registros.append({
            "Codigo"             : cod,
            "Nome"               : nome,
            "Grupo_Sinal"        : grupos[cod],
            "Cluster_ID"         : cluster_id,
            "Membros_Cluster"    : len(membros_cl),
            "N_Lags"             : n_lags,
            "LB_Pvalor"          : round(acf_info["lb_pvalor"], 4)
                                     if not np.isnan(acf_info["lb_pvalor"]) else np.nan,
            # Pooled models
            "Pooled_Ridge_RMSE"  : round(r_ridge["RMSE"], 2),
            "Pooled_Ridge_MAE"   : round(r_ridge["MAE"],  2),
            "Pooled_Ridge_MAPE"  : round(r_ridge["MAPE"], 4),
            "Pooled_Ridge_Params": r_ridge["params"],
            "Pooled_SVR_RMSE"    : round(r_svr["RMSE"],   2),
            "Pooled_SVR_MAE"     : round(r_svr["MAE"],    2),
            "Pooled_SVR_MAPE"    : round(r_svr["MAPE"],   4),
            "Pooled_SVR_Params"  : r_svr["params"],
            "Pooled_MLP_RMSE"    : round(r_mlp["RMSE"],      2),
            "Pooled_MLP_MAE"     : round(r_mlp["MAE"],       2),
            "Pooled_MLP_MAPE"    : round(r_mlp["MAPE"],      4),
            "Pooled_MLP_Params"  : r_mlp["params"],
            # LGBM cluster
            "LGBM_Cluster_RMSE"  : round(r_lgbm_cl["RMSE"],  2) if not np.isnan(r_lgbm_cl["RMSE"]) else np.nan,
            "LGBM_Cluster_MAE"   : round(r_lgbm_cl["MAE"],   2) if not np.isnan(r_lgbm_cl["MAE"])  else np.nan,
            "LGBM_Cluster_MAPE"  : round(r_lgbm_cl["MAPE"],  4) if not np.isnan(r_lgbm_cl["MAPE"]) else np.nan,
            "LGBM_Cluster_Params": r_lgbm_cl.get("params", ""),
        })
        _salvar_checkpoint_local(registros, timestamp, concluidos)

        # Gráfico de previsão (se habilitado)
        if gerar_graficos:
            try:
                rmses_pool = {
                    "Ridge": r_ridge["RMSE"],
                    "SVR"  : r_svr["RMSE"],
                    "MLP"  : r_mlp["RMSE"],
                }
                melhor_modelo_local = min(
                    (m for m, v in rmses_pool.items() if not np.isnan(v)),
                    key=rmses_pool.get,
                    default="Ridge",
                )
                resultado_melhor = {
                    "Ridge": r_ridge, "SVR": r_svr, "MLP": r_mlp,
                }[melhor_modelo_local]
                resultado_melhor_plot = dict(resultado_melhor)
                resultado_melhor_plot["preds_teste"] = resultado_melhor.get(
                    "preds_teste", resultado_melhor.get("preds", [])
                )
                meses_prev_period = [pd.Period(m, freq="M") for m in MESES_PREV]
                plotar_serie(
                    codigo      = str(cod),
                    nome        = nome,
                    grupo_sinal = grupos[cod],
                    treino_idx  = treino_indices[cod],
                    treino_vals = treinos[cod],
                    teste_idx   = teste_indices[cod],
                    teste_vals  = teste_vals,
                    resultados_serie = {
                        "Pooled_" + melhor_modelo_local: resultado_melhor_plot
                    },
                    melhor_modelo = "Pooled_" + melhor_modelo_local,
                    meses_prev  = meses_prev_period,
                    fc_vals     = [],
                    n_holdout   = len(teste_vals),
                    n_lags      = n_lags,
                    mostrar     = False,
                    sufixo      = "_local",
                )
            except Exception as exc_plot:
                logger.warning("[%s] Erro ao gerar gráfico local: %s", cod, exc_plot)

    comparacao_df = pd.DataFrame(registros)

    # 5. Junta com resultados individuais
    ind_df = _carregar_resultados_individuais()

    if ind_df is not None:
        # Seleciona apenas o modelo vencedor de cada serie (Selecionado == True)
        ind_sel = (
            ind_df[ind_df["Selecionado"] == True]
            [[
                "Codigo", "Modelo", "RMSE", "MAE", "MAPE",
                "N_Treino", "N_Holdout", "N_Lags",
            ]]
            .rename(columns={
                "Modelo"   : "Ind_Melhor_Modelo",
                "RMSE"     : "Ind_RMSE",
                "MAE"      : "Ind_MAE",
                "MAPE"     : "Ind_MAPE",
                "N_Treino" : "N_Treino",
                "N_Holdout": "N_Holdout",
            })
            .drop_duplicates("Codigo")
            .assign(Codigo=lambda x: x["Codigo"].astype(str))
        )
        comparacao_df["Codigo"] = comparacao_df["Codigo"].astype(str)
        comparacao_df = comparacao_df.merge(
            ind_sel, on="Codigo", how="left"
        )

        # Melhor pooled (menor RMSE entre Ridge/SVR/MLP)
        rmse_pooled = comparacao_df[[
            "Pooled_Ridge_RMSE", "Pooled_SVR_RMSE", "Pooled_MLP_RMSE"
        ]].values
        modelos_p = ["Ridge", "SVR", "MLP"]
        best_col  = np.nanargmin(rmse_pooled, axis=1)
        comparacao_df["Pooled_Melhor_Modelo"] = [modelos_p[i] for i in best_col]
        comparacao_df["Pooled_Melhor_RMSE"]   = np.nanmin(rmse_pooled, axis=1)

        # Delta: (RMSE_pooled - RMSE_ind) / RMSE_ind * 100
        comparacao_df["Delta_RMSE_pct"] = (
            (comparacao_df["Pooled_Melhor_RMSE"] - comparacao_df["Ind_RMSE"])
            / comparacao_df["Ind_RMSE"].abs().replace(0, np.nan)
            * 100
        ).round(2)

        # 6. Teste de Wilcoxon (ind_RMSE vs melhor_pooled_RMSE)
        arr_ind    = comparacao_df["Ind_RMSE"].values.astype(float)
        arr_pooled = comparacao_df["Pooled_Melhor_RMSE"].values.astype(float)
        stat_w, p_w = _wilcoxon_safe(arr_ind, arr_pooled)

        if verbose:
            print("\n" + "=" * 60)
            print("COMPARACAO: Modelos Individuais vs Pooled (Local)")
            print("=" * 60)
            for _, row in comparacao_df.iterrows():
                print(f"\n  {row['Codigo']}  {str(row['Nome'])[:45]}")
                print(f"    Individual ({row.get('Ind_Melhor_Modelo','N/A')}): "
                      f"RMSE={row.get('Ind_RMSE', float('nan')):>14,.2f}")
                print(f"    Pooled melhor ({row['Pooled_Melhor_Modelo']}): "
                      f"RMSE={row['Pooled_Melhor_RMSE']:>14,.2f}  "
                      f"delta={row['Delta_RMSE_pct']:+.1f}%")

            print("\n" + "-" * 60)
            print(f"RMSE medio -- Individual : "
                  f"{comparacao_df['Ind_RMSE'].mean():>16,.2f}")
            print(f"RMSE medio -- Ridge pool : "
                  f"{comparacao_df['Pooled_Ridge_RMSE'].mean():>16,.2f}")
            print(f"RMSE medio -- SVR pool   : "
                  f"{comparacao_df['Pooled_SVR_RMSE'].mean():>16,.2f}")
            print(f"RMSE medio -- MLP pool   : "
                  f"{comparacao_df['Pooled_MLP_RMSE'].mean():>16,.2f}")
            print(f"RMSE medio -- Melhor pool: "
                  f"{comparacao_df['Pooled_Melhor_RMSE'].mean():>16,.2f}")
            print(f"\nDelta medio (pool vs ind): "
                  f"{comparacao_df['Delta_RMSE_pct'].mean():+.2f}%  "
                  f"(negativo = pool melhor)")
            if not np.isnan(p_w):
                print(f"Wilcoxon signed-rank: stat={stat_w:.2f}  p={p_w:.4f}  "
                      f"({'diferenca significativa' if p_w < 0.05 else 'sem diferenca significativa'} "
                      f"a 5%)")
            else:
                print("Wilcoxon: amostras insuficientes (n < 5)")
    else:
        stat_w = float("nan")
        p_w    = float("nan")

    # 7. Salva resultados
    if _CHECKPOINT_PATH.exists():
        _CHECKPOINT_PATH.unlink()
        logger.info("Checkpoint local removido após conclusão bem-sucedida.")

    csv_comp = RESULTS_DIR / f"comparacao_local_{timestamp}.csv"
    comparacao_df.to_csv(csv_comp, index=False)
    logger.info("CSV de comparacao salvo: %s", csv_comp.name)

    # Excel de comparacao
    _gerar_excel_comparacao(comparacao_df, timestamp)

    return {
        "comparacao_df"      : comparacao_df,
        "clustering"         : resultado_clustering,
        "features_df"        : features_df,
        "timestamp"          : timestamp,
        "wilcoxon_stat"      : stat_w,
        "wilcoxon_p"         : p_w,
    }


# Excel de comparacao

def _gerar_excel_comparacao(df: pd.DataFrame, timestamp: str) -> Path:
    """
    Gera Excel formatado com a comparacao individual vs pooled.
    Colunas com melhor modelo por linha destacadas em verde.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _H_BG   = "1B3A5C"; _H_FG = "FFFFFF"
    _ROW_A  = "FFFFFF"; _ROW_B = "EBF3FB"
    _GRN_BG = "D5F5E3"; _GRN_FG = "1E8449"
    _RED_BG = "FDECEA"; _RED_FG = "C0392B"
    _TOT_BG = "D6E8F7"; _TOT_FG = "1B3A5C"
    _FONT   = "Arial"

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10):
        return Font(name=_FONT, bold=bold, color=color, size=size)
    def _border():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    headers = [
        "Codigo", "Nome", "Grupo Sinal", "Cluster",
        "Membros\nCluster", "N Lags",
        # Individual
        "Ind.\nModelo", "Ind.\nRMSE", "Ind.\nMAPE %",
        # Pooled por modelo
        "Ridge\nRMSE", "Ridge\nMAPE %",
        "SVR\nRMSE",   "SVR\nMAPE %",
        "MLP\nRMSE",   "MLP\nMAPE %",
        # Melhor pooled
        "Melhor\nPooled", "Melhor\nPooled RMSE",
        # Delta
        "Delta RMSE\n(pool vs ind) %",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparacao"

    # Cabecalho
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(_H_BG); cell.font = _font(True, _H_FG)
        cell.alignment = _center(); cell.border = _border()

    for r_idx, row in df.iterrows():
        er   = r_idx + 2
        bg   = _ROW_B if r_idx % 2 == 0 else _ROW_A
        delta = row.get("Delta_RMSE_pct", float("nan"))

        valores = [
            row.get("Codigo", ""),
            row.get("Nome", ""),
            row.get("Grupo_Sinal", ""),
            row.get("Cluster_ID", ""),
            row.get("Membros_Cluster", ""),
            row.get("N_Lags", ""),
            row.get("Ind_Melhor_Modelo", ""),
            row.get("Ind_RMSE", np.nan),
            row.get("Ind_MAPE", np.nan),
            row.get("Pooled_Ridge_RMSE", np.nan),
            row.get("Pooled_Ridge_MAPE", np.nan),
            row.get("Pooled_SVR_RMSE",   np.nan),
            row.get("Pooled_SVR_MAPE",   np.nan),
            row.get("Pooled_MLP_RMSE",   np.nan),
            row.get("Pooled_MLP_MAPE",   np.nan),
            row.get("Pooled_Melhor_Modelo", ""),
            row.get("Pooled_Melhor_RMSE",  np.nan),
            delta,
        ]

        for c_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.border = _border()

            # Delta: verde se pool melhor (< 0), vermelho se pior (> 0)
            if c_idx == 18 and isinstance(val, float) and not np.isnan(val):
                cell.fill = _fill(_GRN_BG if val < 0 else _RED_BG)
                cell.font = _font(True, _GRN_FG if val < 0 else _RED_FG)
                cell.number_format = '+0.00%;-0.00%'
                cell.value = val / 100
                cell.alignment = _center()
            elif c_idx in (8, 10, 12, 14, 17):  # RMSEs
                cell.fill = _fill(bg); cell.font = _font()
                cell.number_format = '#,##0.00'; cell.alignment = _center()
            elif c_idx in (9, 11, 13, 15):  # MAPEs
                cell.fill = _fill(bg); cell.font = _font()
                cell.number_format = '0.00"%"'; cell.alignment = _center()
            else:
                cell.fill = _fill(bg); cell.font = _font()
                cell.alignment = _center() if c_idx in (1,4,5,6) else _left()

    # Rodape com medias
    last = len(df) + 1
    rod  = last + 2
    cell = ws.cell(row=rod, column=1, value=f"Media -- {len(df)} series")
    cell.fill = _fill(_TOT_BG); cell.font = _font(True, _TOT_FG)
    cell.border = _border(); cell.alignment = _left()
    for c in range(2, len(headers) + 1):
        cl = ws.cell(row=rod, column=c)
        cl.fill = _fill(_TOT_BG); cl.border = _border(); cl.alignment = _center()
    for c_idx in [8, 10, 12, 14, 17]:
        col_l = get_column_letter(c_idx)
        cl = ws.cell(row=rod, column=c_idx,
                     value=f"=AVERAGE({col_l}2:{col_l}{last})")
        cl.fill = _fill(_TOT_BG); cl.font = _font(True, _TOT_FG)
        cl.number_format = '#,##0.00'; cl.border = _border(); cl.alignment = _center()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
    ws.row_dimensions[1].height = 36

    # Ajusta larguras
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value), default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(45, max(8, max_len + 3))

    caminho = RESULTS_DIR / f"comparacao_local_{timestamp}.xlsx"
    wb.save(caminho)
    logger.info("Excel de comparacao salvo: %s", caminho.name)
    return caminho


# Entry point

if __name__ == "__main__":
    resultados = rodar_experimento_local(verbose=True)

    comp = resultados["comparacao_df"]
    print("\n" + "=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    if "Ind_RMSE" in comp.columns and "Pooled_Melhor_RMSE" in comp.columns:
        n_melhor_pool = (comp["Pooled_Melhor_RMSE"] < comp["Ind_RMSE"]).sum()
        print(f"Series onde pooled < individual: {n_melhor_pool} / {len(comp)}")
        print(f"RMSE medio individual : {comp['Ind_RMSE'].mean():>16,.2f}")
        print(f"RMSE medio pooled     : {comp['Pooled_Melhor_RMSE'].mean():>16,.2f}")
