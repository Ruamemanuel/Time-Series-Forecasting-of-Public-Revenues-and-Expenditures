"""
Gera dois Excel ao fim do experimento:

  1. diagnostico_series_<timestamp>.xlsx — uma linha por série, com o diagnóstico
     estatístico (Ljung-Box + PACF).
  2. previsoes_<timestamp>.xlsx — uma linha por série prevista, com o modelo
     vencedor e os primeiros meses de previsão.
"""

from __future__ import annotations

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import RESULTS_DIR

logger = logging.getLogger(__name__)

# Paleta
_H_BG    = "1B3A5C"
_H_FG    = "FFFFFF"
_ROW_A   = "FFFFFF"
_ROW_B   = "EBF3FB"
_RB_BG   = "FDECEA"
_RB_FG   = "C0392B"
_TOT_BG  = "D6E8F7"
_TOT_FG  = "1B3A5C"
_GRN     = "196F3D"
_FONT    = "Arial"

_COR_MODELO = {
    "ARIMA": "FADBD8",
    "LR":    "FDEBD0",
    "SVR":   "D5F5E3",
    "MLP":   "E8DAEF",
}
_TXT_MODELO = {
    "ARIMA": "C0392B",
    "LR":    "CA6F1E",
    "SVR":   "1E8449",
    "MLP":   "6C3483",
}


# Helpers de estilo

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name=_FONT, bold=bold, color=color, size=size)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _thin_border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _aplicar_cabecalho(ws, headers: list[str]) -> None:
    for col, titulo in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill      = _fill(_H_BG)
        cell.font      = _font(bold=True, color=_H_FG, size=10)
        cell.alignment = _center()
        cell.border    = _thin_border()

def _ajustar_colunas(ws, min_w=8, max_w=45) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 3))

def _rodape_label(ws, row: int, n_cols: int, texto: str) -> None:
    cell = ws.cell(row=row, column=1, value=texto)
    cell.font      = _font(bold=True, color=_TOT_FG, size=9)
    cell.fill      = _fill(_TOT_BG)
    cell.alignment = _left()
    cell.border    = _thin_border()
    for col in range(2, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(_TOT_BG)
        c.border = _thin_border()


# Excel 1: Diagnóstico de séries

def gerar_excel_diagnostico(
    metricas_df: pd.DataFrame,
    timestamp: str,
) -> Path:
    """Cria diagnostico_series_<timestamp>.xlsx com diagnóstico LB + PACF."""
    diag = (
        metricas_df
        .drop_duplicates(subset="Codigo", keep="first")
        [["Codigo", "Nome", "Grupo_Sinal", "N_Total", "N_Treino", "N_Holdout",
          "N_Lags", "LB_Pvalor", "Lags_Sig_PACF", "Metodo_Lags",
          "Ruido_Branco"]]
        .sort_values(["Ruido_Branco", "Codigo"], ascending=[True, True])
        .reset_index(drop=True)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnóstico"

    headers = [
        "Código Ação", "Nome da Ação", "Grupo de Sinal",
        "N Total", "N Treino", "N Conjunto de teste",
        "N Lags\nSelecionado", "LB p-valor", "Conclusão LB",
        "Lags PACF\nSignificativos", "Método Seleção\nde Lags",
        "Ruído Branco?",
    ]
    _aplicar_cabecalho(ws, headers)

    for r_idx, row in diag.iterrows():
        excel_row = r_idx + 2
        is_rb     = bool(row.get("Ruido_Branco", False))
        lb_p      = row.get("LB_Pvalor", np.nan)

        if pd.isna(lb_p):
            conclusao_lb = "Indisponível"
        elif is_rb:
            conclusao_lb = f"Não rejeita H₀ (p={lb_p:.4f})"
        else:
            conclusao_lb = f"Rejeita H₀ (p={lb_p:.4f})"

        valores = [
            row.get("Codigo", ""),
            row.get("Nome", ""),
            row.get("Grupo_Sinal", ""),
            row.get("N_Total", ""),
            row.get("N_Treino", ""),
            row.get("N_Holdout", ""),
            row.get("N_Lags", 0) if not is_rb else 0,
            lb_p if not pd.isna(lb_p) else "N/D",
            conclusao_lb,
            row.get("Lags_Sig_PACF", "[]"),
            row.get("Metodo_Lags", ""),
            "Sim" if is_rb else "Não",
        ]

        bg = _RB_BG if is_rb else (_ROW_B if r_idx % 2 == 0 else _ROW_A)
        fg = _RB_FG if is_rb else "000000"

        for c_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=is_rb, color=fg)
            cell.border    = _thin_border()
            cell.alignment = _center() if c_idx in (1, 4, 5, 6, 7, 8, 12) else _left()
            if c_idx == 8 and isinstance(val, float):
                cell.number_format = "0.0000"

    last_data_row = len(diag) + 1
    rodape_row    = last_data_row + 2
    n_rb  = int(diag["Ruido_Branco"].sum())
    n_tot = len(diag)
    _rodape_label(ws, rodape_row, len(headers),
                  f"Total de séries: {n_tot}  |  Ruído branco: {n_rb}  |  "
                  f"Previstas: {n_tot - n_rb}")

    leg_row = rodape_row + 2
    ws.cell(row=leg_row, column=1, value="Legenda:").font = _font(bold=True)
    c1 = ws.cell(row=leg_row, column=2, value="Série prevista")
    c1.fill = _fill(_ROW_B); c1.font = _font(); c1.border = _thin_border()
    c2 = ws.cell(row=leg_row, column=3, value="Ruído branco — série pulada")
    c2.fill = _fill(_RB_BG); c2.font = _font(bold=True, color=_RB_FG); c2.border = _thin_border()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data_row}"
    _ajustar_colunas(ws)
    ws.row_dimensions[1].height = 32

    caminho = RESULTS_DIR / f"diagnostico_series_{timestamp}.xlsx"
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    logger.info("Excel diagnóstico salvo: %s", caminho.name)
    return caminho


# Excel 2: Previsões

def gerar_excel_previsoes(
    prev_df: pd.DataFrame,
    timestamp: str,
) -> Path:
    """Cria previsoes_<timestamp>.xlsx com o modelo escolhido e as previsões
    mensais. As colunas Prev_YYYY_MM e Var_YYYY_MM_pct são detectadas
    automaticamente no DataFrame."""
    df = prev_df.sort_values("Codigo").reset_index(drop=True)

    # Detecta colunas dinâmicas de previsão e variação
    prev_cols = sorted([c for c in df.columns if c.startswith("Prev_")])
    var_cols  = sorted([c for c in df.columns if c.startswith("Var_") and c.endswith("_pct")])
    # Coluna do último real (Ultimo_Real_YYYY_MM)
    real_cols = [c for c in df.columns if c.startswith("Ultimo_Real_")]
    real_col  = real_cols[0] if real_cols else None

    # Cabeçalhos de mês legíveis: YYYY_MM → YYYY-MM
    def _label_mes(col_key: str) -> str:
        parts = col_key.split("_")
        return f"{parts[-2]}-{parts[-1]}"

    headers_fixos = [
        "Código Ação", "Nome da Ação", "Grupo de Sinal",
        "Modelo\nSelecionado", "RMSE\n(conjunto de teste)", "MAPE %\n(conjunto de teste)",
        "LB p-valor", "Método Lags",
        f"Último Real\n{_label_mes(real_col) if real_col else ''} (R$)",
    ]
    headers_prev = [f"Prev.\n{_label_mes(c)} (R$)" for c in prev_cols]
    headers_var  = [f"Var.\n{_label_mes(c)} (%)" for c in var_cols]
    headers = headers_fixos + headers_prev + headers_var

    n_fixos   = len(headers_fixos)
    n_prev    = len(prev_cols)
    # Índices 1-based das colunas de previsão e variação no Excel
    idx_prev_start = n_fixos + 1
    idx_var_start  = n_fixos + n_prev + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Previsões"
    _aplicar_cabecalho(ws, headers)

    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2
        modelo    = str(row.get("Melhor_Modelo", ""))
        lb_p      = row.get("LB_Pvalor", np.nan)
        bg_base   = _ROW_B if r_idx % 2 == 0 else _ROW_A

        valores_fixos = [
            row.get("Codigo", ""),
            row.get("Nome", ""),
            row.get("Grupo_Sinal", ""),
            modelo,
            row.get("RMSE_Teste", np.nan),
            row.get("MAPE_Teste_Pct", np.nan),
            lb_p if not (isinstance(lb_p, float) and np.isnan(lb_p)) else "N/D",
            row.get("Metodo_Lags", ""),
            row.get(real_col, np.nan) if real_col else np.nan,
        ]
        valores_prev = [row.get(c, np.nan) for c in prev_cols]
        valores_var  = [row.get(c, np.nan) for c in var_cols]
        valores = valores_fixos + valores_prev + valores_var

        for c_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.border = _thin_border()

            if c_idx == 4:
                bg  = _COR_MODELO.get(modelo, bg_base)
                fg  = _TXT_MODELO.get(modelo, "000000")
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg)
                cell.alignment = _center()
            else:
                cell.fill = _fill(bg_base)
                cell.font = _font()
                cell.alignment = _center() if c_idx in (1, 5, 6, 7, 8) else _left()

            # Formatos numéricos
            if c_idx == 5:
                cell.number_format = '#,##0.00'
            elif c_idx == 6:
                cell.number_format = '0.00%'
                if isinstance(val, float) and not np.isnan(val):
                    cell.value = val / 100
            elif c_idx == 7 and isinstance(val, float):
                cell.number_format = '0.0000'
            elif idx_prev_start <= c_idx < idx_var_start:
                cell.number_format = '#,##0.00'
                cell.alignment = _center()
            elif c_idx >= idx_var_start:
                cell.number_format = '0.00"%"'
                cell.alignment = _center()
                if isinstance(val, float) and not np.isnan(val):
                    cell.font = _font(color=(_GRN if val >= 0 else _RB_FG))

    last_data = len(df) + 1
    rodape    = last_data + 2
    _rodape_label(ws, rodape, len(headers), f"Média — {len(df)} séries previstas")

    # Médias para RMSE, MAPE e variações mensais
    for c_idx in [5, 6] + list(range(idx_var_start, len(headers) + 1)):
        col_l = get_column_letter(c_idx)
        cell  = ws.cell(row=rodape, column=c_idx,
                        value=f"=AVERAGE({col_l}2:{col_l}{last_data})")
        cell.fill      = _fill(_TOT_BG)
        cell.font      = _font(bold=True, color=_TOT_FG)
        cell.border    = _thin_border()
        cell.alignment = _center()
        cell.number_format = '#,##0.00' if c_idx == 5 else ('0.00%' if c_idx == 6 else '0.00"%"')

    leg_row = rodape + 2
    ws.cell(row=leg_row, column=1, value="Legenda modelos:").font = _font(bold=True)
    for i, (mod, bg) in enumerate(_COR_MODELO.items(), 2):
        c = ws.cell(row=leg_row, column=i, value=mod)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, color=_TXT_MODELO[mod])
        c.border    = _thin_border()
        c.alignment = _center()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data}"
    _ajustar_colunas(ws)
    ws.row_dimensions[1].height = 36

    caminho = RESULTS_DIR / f"previsoes_{timestamp}.xlsx"
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    logger.info("Excel previsões salvo: %s", caminho.name)
    return caminho


# Orquestrador

def gerar_excel_resultados(
    metricas_df: pd.DataFrame,
    prev_df: pd.DataFrame,
    timestamp: str,
) -> dict:
    """Gera os dois arquivos Excel e retorna os caminhos."""
    path_diag = gerar_excel_diagnostico(metricas_df, timestamp)
    path_prev = gerar_excel_previsoes(prev_df, timestamp)
    return {"diagnostico": path_diag, "previsoes": path_prev}
