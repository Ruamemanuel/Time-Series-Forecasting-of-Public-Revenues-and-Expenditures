"""
Monta a tabela unificada comparando todos os paradigmas por série: o melhor
individual (de metricas_*.csv), os pooled e o LGBM-cluster (de
comparacao_local_*.csv), os globais Ridge/RF/MLP/LGBM/Chronos (de
comparacao_global_*.csv), mais os "melhores" de cada grupo e um oráculo (o menor
RMSE entre todos os não-individuais).

Para cada paradigma calcula RMSE e MAPE (médio e mediano), as vitórias contra o
individual, o delta percentual mediano e o p-valor do Wilcoxon bilateral. Salva
results/comparacao_unificada_<timestamp>.csv e um .xlsx com três abas (Resumo,
Completo, Vitórias).

Rodar com: python comparar_todos.py  (ou python run_all.py --compare)
"""

from __future__ import annotations

import glob
import logging
import sys
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import wilcoxon

warnings.filterwarnings("ignore")

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

from config import RESULTS_DIR

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s -- %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("comparar_todos")

# Definicao dos paradigmas  {rotulo: coluna_RMSE}

PARADIGMAS_RMSE: dict[str, str] = {
    "Individual"    : "Ind_RMSE",
    "Pooled Ridge"  : "Pooled_Ridge_RMSE",
    "Pooled SVR"    : "Pooled_SVR_RMSE",
    "Pooled MLP"    : "Pooled_MLP_RMSE",
    "Pooled Melhor" : "Pooled_Melhor_RMSE",
    "LGBM Cluster"  : "LGBM_Cluster_RMSE",
    "Global Ridge"  : "Global_Ridge_RMSE",
    "Global RF"     : "Global_RF_RMSE",
    "Global MLP"    : "Global_MLP_RMSE",
    "Global LGBM"   : "Global_LGBM_RMSE",
    "Global Melhor" : "Global_Melhor_RMSE",
    "Global Chronos": "Global_Chronos_RMSE",
    "Oracle"        : "_Oracle_RMSE",        # computado abaixo
}

PARADIGMAS_MAPE: dict[str, str] = {
    "Individual"    : "Ind_MAPE",
    "Pooled Ridge"  : "Pooled_Ridge_MAPE",
    "Pooled SVR"    : "Pooled_SVR_MAPE",
    "Pooled MLP"    : "Pooled_MLP_MAPE",
    "Pooled Melhor" : "_Pooled_Melhor_MAPE",  # computado abaixo
    "LGBM Cluster"  : "LGBM_Cluster_MAPE",
    "Global Ridge"  : "Global_Ridge_MAPE",
    "Global RF"     : "Global_RF_MAPE",
    "Global MLP"    : "Global_MLP_MAPE",
    "Global LGBM"   : "Global_LGBM_MAPE",
    "Global Melhor" : "_Global_Melhor_MAPE",  # computado abaixo
    "Global Chronos": "Global_Chronos_MAPE",
    "Oracle"        : "_Oracle_MAPE",          # computado abaixo
}


# Helpers

def _mais_recente(padrao: str) -> str | None:
    """Retorna o arquivo mais recente que bate com o padrao (glob)."""
    arquivos = sorted(glob.glob(padrao))
    return arquivos[-1] if arquivos else None


def _wilcoxon_safe(a: np.ndarray, b: np.ndarray) -> float:
    """Wilcoxon signed-rank test bilateral; retorna p-valor ou nan."""
    diff = a - b
    diff = diff[~np.isnan(diff)]
    if len(diff) < 5:
        return float("nan")
    try:
        _, p = wilcoxon(diff, alternative="two-sided", zero_method="wilcox")
        return float(p)
    except Exception:
        return float("nan")


def _carregar_metricas() -> pd.DataFrame | None:
    """
    Carrega o CSV de metricas individuais (experiment.py) e retorna
    uma linha por serie com o melhor modelo selecionado.
    """
    padrao = str(RESULTS_DIR / "metricas_*.csv")
    # Exclui arquivos de comparacao que contem 'metricas' acidentalmente
    candidatos = [
        c for c in sorted(glob.glob(padrao))
        if "comparacao" not in Path(c).name
    ]
    if not candidatos:
        return None
    path = candidatos[-1]
    logger.info("Metricas individuais: %s", Path(path).name)
    df = pd.read_csv(path)

    # Filtra apenas o modelo selecionado por serie
    if "Selecionado" in df.columns:
        df_sel = df[df["Selecionado"] == True].copy()
    else:
        # Fallback: menor RMSE por serie
        df_sel = df.loc[df.groupby("Codigo")["RMSE"].idxmin()].copy()

    df_sel = df_sel.rename(columns={
        "Modelo" : "Ind_Melhor_Modelo",
        "RMSE"   : "Ind_RMSE",
        "MAE"    : "Ind_MAE",
        "MAPE"   : "Ind_MAPE",
    })
    cols = ["Codigo", "Nome", "Ind_Melhor_Modelo", "Ind_RMSE", "Ind_MAE", "Ind_MAPE"]
    cols = [c for c in cols if c in df_sel.columns]
    df_sel["Codigo"] = df_sel["Codigo"].astype(str)
    return df_sel[cols].drop_duplicates("Codigo")


def _carregar_local() -> pd.DataFrame | None:
    """Carrega o CSV mais recente do experimento local (pooled + LGBM cluster)."""
    path = _mais_recente(str(RESULTS_DIR / "comparacao_local_*.csv"))
    if not path:
        return None
    logger.info("Resultados locais: %s", Path(path).name)
    df = pd.read_csv(path)
    df["Codigo"] = df["Codigo"].astype(str)
    return df


def _carregar_global() -> pd.DataFrame | None:
    """Carrega o CSV mais recente do experimento global."""
    path = _mais_recente(str(RESULTS_DIR / "comparacao_global_*.csv"))
    if not path:
        return None
    logger.info("Resultados globais: %s", Path(path).name)
    df = pd.read_csv(path)
    df["Codigo"] = df["Codigo"].astype(str)
    return df


# Montagem do DataFrame unificado

def _montar_unificado(
    df_met   : pd.DataFrame | None,
    df_local : pd.DataFrame | None,
    df_glob  : pd.DataFrame | None,
) -> pd.DataFrame:
    """
    Combina as tres fontes num unico DataFrame indexado por Codigo.
    Fallbacks graciosamente para fontes ausentes.
    """
    # Parte base: serie + metricas individuais
    if df_met is not None:
        base = df_met.copy()
    elif df_local is not None:
        cols_ind = ["Codigo", "Nome", "Ind_Melhor_Modelo",
                    "Ind_RMSE", "Ind_MAE", "Ind_MAPE"]
        cols_ind = [c for c in cols_ind if c in df_local.columns]
        base = df_local[cols_ind].drop_duplicates("Codigo").copy()
    elif df_glob is not None:
        cols_ind = ["Codigo", "Nome", "Ind_Melhor_Modelo",
                    "Ind_RMSE", "Ind_MAE", "Ind_MAPE"]
        cols_ind = [c for c in cols_ind if c in df_glob.columns]
        base = df_glob[cols_ind].drop_duplicates("Codigo").copy()
    else:
        logger.error("Nenhuma fonte de dados encontrada.")
        return pd.DataFrame()

    # Merge com resultados locais
    if df_local is not None:
        colunas_locais = [
            "Codigo",
            "Grupo_Sinal", "Cluster_ID",
            "Pooled_Ridge_RMSE", "Pooled_Ridge_MAE", "Pooled_Ridge_MAPE",
            "Pooled_SVR_RMSE",   "Pooled_SVR_MAE",   "Pooled_SVR_MAPE",
            "Pooled_MLP_RMSE",   "Pooled_MLP_MAE",   "Pooled_MLP_MAPE",
            "Pooled_Melhor_Modelo", "Pooled_Melhor_RMSE",
            "LGBM_Cluster_RMSE", "LGBM_Cluster_MAE", "LGBM_Cluster_MAPE",
        ]
        colunas_locais = [c for c in colunas_locais if c in df_local.columns]
        base = base.merge(
            df_local[colunas_locais].drop_duplicates("Codigo"),
            on="Codigo", how="left",
        )

    # Merge com resultados globais
    if df_glob is not None:
        colunas_globais = [
            "Codigo",
            "N_Treino", "N_Holdout",
            "Global_Ridge_RMSE", "Global_Ridge_MAE", "Global_Ridge_MAPE",
            "Global_RF_RMSE",    "Global_RF_MAE",    "Global_RF_MAPE",
            "Global_MLP_RMSE",   "Global_MLP_MAE",   "Global_MLP_MAPE",
            "Global_LGBM_RMSE",    "Global_LGBM_MAE",    "Global_LGBM_MAPE",
            "Global_Chronos_RMSE", "Global_Chronos_MAE", "Global_Chronos_MAPE",
            "Global_Melhor_Modelo", "Global_Melhor_RMSE",
        ]
        colunas_globais = [c for c in colunas_globais if c in df_glob.columns]
        base = base.merge(
            df_glob[colunas_globais].drop_duplicates("Codigo"),
            on="Codigo", how="left",
        )

    return base


def _adicionar_oracle_e_derivados(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adiciona colunas calculadas:
      - _Pooled_Melhor_MAPE : MAPE do melhor modelo pooled (por serie)
      - _Global_Melhor_MAPE : MAPE do melhor modelo global (por serie)
      - _Oracle_RMSE        : min de todos os paradigmas nao-individuais
      - _Oracle_MAPE        : MAPE correspondente ao oracle de RMSE
    """
    # MAPE do melhor modelo pooled
    colunas_mape_pooled = ["Pooled_Ridge_MAPE", "Pooled_SVR_MAPE", "Pooled_MLP_MAPE",
                            "LGBM_Cluster_MAPE"]
    colunas_mape_pooled = [c for c in colunas_mape_pooled if c in df.columns]
    colunas_rmse_pooled = ["Pooled_Ridge_RMSE", "Pooled_SVR_RMSE", "Pooled_MLP_RMSE",
                            "LGBM_Cluster_RMSE"]
    colunas_rmse_pooled = [c for c in colunas_rmse_pooled if c in df.columns]

    if colunas_rmse_pooled and colunas_mape_pooled:
        mat_rmse = df[colunas_rmse_pooled].values.astype(float)
        mat_mape = df[colunas_mape_pooled].values.astype(float)
        best_idx = np.nanargmin(mat_rmse, axis=1) if len(colunas_rmse_pooled) else None
        if best_idx is not None:
            df["_Pooled_Melhor_MAPE"] = [
                mat_mape[i, best_idx[i]] if not np.isnan(mat_rmse[i, best_idx[i]])
                else float("nan")
                for i in range(len(df))
            ]

    # MAPE do melhor modelo global (sem Chronos — mantém compatibilidade)
    colunas_mape_glob = ["Global_Ridge_MAPE", "Global_RF_MAPE",
                          "Global_MLP_MAPE",   "Global_LGBM_MAPE"]
    colunas_mape_glob = [c for c in colunas_mape_glob if c in df.columns]
    colunas_rmse_glob = ["Global_Ridge_RMSE", "Global_RF_RMSE",
                          "Global_MLP_RMSE",   "Global_LGBM_RMSE"]
    colunas_rmse_glob = [c for c in colunas_rmse_glob if c in df.columns]

    if colunas_rmse_glob and colunas_mape_glob:
        mat_rmse = df[colunas_rmse_glob].values.astype(float)
        mat_mape = df[colunas_mape_glob].values.astype(float)
        best_idx = np.nanargmin(mat_rmse, axis=1) if len(colunas_rmse_glob) else None
        if best_idx is not None:
            df["_Global_Melhor_MAPE"] = [
                mat_mape[i, best_idx[i]] if not np.isnan(mat_rmse[i, best_idx[i]])
                else float("nan")
                for i in range(len(df))
            ]

    # Recalcula Global_Melhor incluindo Chronos
    _modelos_glob = {
        "Ridge"  : ("Global_Ridge_RMSE",   "Global_Ridge_MAPE"),
        "RF"     : ("Global_RF_RMSE",      "Global_RF_MAPE"),
        "MLP"    : ("Global_MLP_RMSE",     "Global_MLP_MAPE"),
        "LGBM"   : ("Global_LGBM_RMSE",    "Global_LGBM_MAPE"),
        "Chronos": ("Global_Chronos_RMSE", "Global_Chronos_MAPE"),
    }
    _modelos_glob = {k: v for k, v in _modelos_glob.items() if v[0] in df.columns}

    if _modelos_glob:
        nomes_mod  = list(_modelos_glob.keys())
        cols_rmse  = [_modelos_glob[m][0] for m in nomes_mod]
        cols_mape  = [_modelos_glob[m][1] for m in nomes_mod]
        mat_r = df[cols_rmse].values.astype(float)
        best  = np.nanargmin(mat_r, axis=1)
        df["Global_Melhor_Modelo"] = [nomes_mod[b] for b in best]
        df["Global_Melhor_RMSE"]   = [mat_r[i, best[i]] for i in range(len(df))]
        cols_mape_pres = [c for c in cols_mape if c in df.columns]
        if len(cols_mape_pres) == len(cols_mape):
            mat_m = df[cols_mape].values.astype(float)
            df["_Global_Melhor_com_Chronos_MAPE"] = [mat_m[i, best[i]] for i in range(len(df))]

    # Oracle: menor RMSE entre todos os paradigmas nao-individuais
    colunas_oracle_rmse = [
        "Pooled_Ridge_RMSE", "Pooled_SVR_RMSE", "Pooled_MLP_RMSE",
        "LGBM_Cluster_RMSE",
        "Global_Ridge_RMSE", "Global_RF_RMSE", "Global_MLP_RMSE", "Global_LGBM_RMSE",
        "Global_Chronos_RMSE",
    ]
    colunas_oracle_rmse = [c for c in colunas_oracle_rmse if c in df.columns]

    colunas_oracle_mape = [c.replace("_RMSE", "_MAPE") for c in colunas_oracle_rmse]
    colunas_oracle_mape = [c for c in colunas_oracle_mape if c in df.columns]

    if colunas_oracle_rmse:
        mat = df[colunas_oracle_rmse].values.astype(float)
        df["_Oracle_RMSE"] = np.nanmin(mat, axis=1)
        if colunas_oracle_mape and len(colunas_oracle_mape) == len(colunas_oracle_rmse):
            mat_mape = df[colunas_oracle_mape].values.astype(float)
            best_idx = np.nanargmin(mat, axis=1)
            df["_Oracle_MAPE"] = [
                mat_mape[i, best_idx[i]] for i in range(len(df))
            ]

    return df


# Estatisticas por paradigma

def _estatisticas_paradigma(
    df         : pd.DataFrame,
    col_rmse   : str,
    col_mape   : str | None,
    ref_rmse   : str = "Ind_RMSE",
    ref_mape   : str = "Ind_MAPE",
) -> dict:
    """
    Calcula estatisticas de um paradigma em relacao ao individual (referencia).
    """
    if col_rmse not in df.columns:
        return {}

    arr_r = df[col_rmse].values.astype(float)
    arr_ind_r = df[ref_rmse].values.astype(float) if ref_rmse in df.columns else None

    mask_r = ~np.isnan(arr_r)
    if arr_ind_r is not None:
        mask_both = mask_r & ~np.isnan(arr_ind_r)
    else:
        mask_both = mask_r

    n_valid = int(mask_both.sum())
    if n_valid == 0:
        return {"N": 0}

    rmse_arr = arr_r[mask_both]
    rmse_med_mean = float(np.mean(rmse_arr))
    rmse_med_medn = float(np.median(rmse_arr))

    stats = {
        "N"            : n_valid,
        "RMSE_Medio"   : rmse_med_mean,
        "RMSE_Mediano" : rmse_med_medn,
    }

    # MAPE
    if col_mape and col_mape in df.columns:
        arr_m = df[col_mape].values.astype(float)
        mask_m = mask_both & ~np.isnan(arr_m)
        if mask_m.sum() > 0:
            stats["MAPE_Medio"]   = float(np.mean(arr_m[mask_m]))
            stats["MAPE_Mediano"] = float(np.median(arr_m[mask_m]))

    if arr_ind_r is not None:
        ind_arr = arr_ind_r[mask_both]
        wins  = int(np.sum(rmse_arr < ind_arr))
        n_ind = n_valid
        stats["Wins"] = wins
        stats["N_Comparavel"] = n_ind
        stats["Win_pct"] = round(100.0 * wins / n_ind, 1) if n_ind > 0 else float("nan")

        delta = (rmse_arr - ind_arr) / np.where(ind_arr == 0, np.nan, np.abs(ind_arr)) * 100
        stats["Delta_Med_pct"]   = float(np.nanmedian(delta))
        stats["Delta_Mean_pct"]  = float(np.nanmean(delta))

        stats["Wilcoxon_p"] = _wilcoxon_safe(ind_arr, rmse_arr)

    return stats


# Main

def comparar_todos(verbose: bool = True) -> pd.DataFrame | None:
    """
    Carrega todas as fontes, monta DataFrame unificado,
    calcula estatisticas e exporta.

    Returns o DataFrame unificado ou None se nenhuma fonte disponivel.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # 1. Carrega fontes
    df_met   = _carregar_metricas()
    df_local = _carregar_local()
    df_glob  = _carregar_global()

    fontes_disponiveis = sum([
        df_met   is not None,
        df_local is not None,
        df_glob  is not None,
    ])

    if fontes_disponiveis == 0:
        logger.error("Nenhuma fonte de dados encontrada em %s", RESULTS_DIR)
        logger.error("Execute experiment.py, experiment_local.py e/ou "
                     "experiment_global.py antes.")
        return None

    logger.info("Fontes disponiveis: %d de 3 (metricas=%s, local=%s, global=%s)",
                fontes_disponiveis,
                df_met   is not None,
                df_local is not None,
                df_glob  is not None)

    # 2. Monta unificado
    df = _montar_unificado(df_met, df_local, df_glob)
    if df.empty:
        return None

    df = _adicionar_oracle_e_derivados(df)

    # 3. Identifica paradigmas disponiveis (coluna presente no df)
    paradigmas_disp = {
        rot: col
        for rot, col in PARADIGMAS_RMSE.items()
        if col in df.columns
    }

    # 4. Calcula estatisticas por paradigma
    resumo_rows = []
    for rotulo, col_rmse in paradigmas_disp.items():
        col_mape = PARADIGMAS_MAPE.get(rotulo)
        stats = _estatisticas_paradigma(df, col_rmse, col_mape)
        if stats:
            row = {"Paradigma": rotulo, **stats}
            resumo_rows.append(row)

    resumo_df = pd.DataFrame(resumo_rows)

    # 5. Exibe tabela de resumo
    if verbose and not resumo_df.empty:
        _exibir_resumo(resumo_df, df)

    # 6. Exporta
    csv_path  = RESULTS_DIR / f"comparacao_unificada_{timestamp}.csv"
    xlsx_path = RESULTS_DIR / f"comparacao_unificada_{timestamp}.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logger.info("CSV salvo: %s", csv_path.name)

    _exportar_excel(df, resumo_df, xlsx_path)
    logger.info("Excel salvo: %s", xlsx_path.name)

    print(f"\nCSV : {csv_path}")
    print(f"XLSX: {xlsx_path}")

    return df


# Exibicao

def _exibir_resumo(resumo_df: pd.DataFrame, df_full: pd.DataFrame) -> None:
    """Imprime a tabela de resumo formatada no console."""
    sep = "=" * 95

    print(f"\n{sep}")
    print("  COMPARACAO UNIFICADA DE PARADIGMAS DE PREVISAO")
    print(f"  Series comparadas: {len(df_full)}")
    print(sep)

    # Cabecalho
    print(f"  {'Paradigma':<18} {'RMSE med':>12} {'RMSE medn':>12} "
          f"{'MAPE med%':>10} {'Wins':>9} {'Win%':>6} "
          f"{'Delta med%':>11} {'Wilcox p':>10}")
    print(f"  {'-'*18} {'-'*12} {'-'*12} {'-'*10} {'-'*9} {'-'*6} "
          f"{'-'*11} {'-'*10}")

    for _, row in resumo_df.iterrows():
        pard    = str(row.get("Paradigma", ""))
        rmse_m  = row.get("RMSE_Medio",   float("nan"))
        rmse_md = row.get("RMSE_Mediano", float("nan"))
        mape_m  = row.get("MAPE_Medio",   float("nan"))
        wins    = row.get("Wins",         float("nan"))
        n_comp  = row.get("N_Comparavel", float("nan"))
        win_pct = row.get("Win_pct",      float("nan"))
        delta   = row.get("Delta_Med_pct", float("nan"))
        p_wx    = row.get("Wilcoxon_p",   float("nan"))

        def _f(v, fmt): return fmt.format(v) if not (isinstance(v, float) and np.isnan(v)) else "  N/A"
        def _fw(w, n):
            if isinstance(w, float) and np.isnan(w): return "  N/A"
            return f"{int(w):>4}/{int(n):<4}"
        def _fp(v):
            if isinstance(v, float) and np.isnan(v): return "  N/A"
            return f"{v:.4f}" if v >= 0.0001 else "<.0001"

        # Linha de referencia (Individual) sem comparacao vs si mesmo
        if pard == "Individual":
            print(f"  {pard:<18} {_f(rmse_m, '{:>12,.0f}')} "
                  f"{_f(rmse_md, '{:>12,.0f}')} "
                  f"{_f(mape_m, '{:>9.1f}%')} "
                  f"{'(ref)':>9} {'':>6} {'':>11} {'':>10}")
        else:
            print(f"  {pard:<18} {_f(rmse_m, '{:>12,.0f}')} "
                  f"{_f(rmse_md, '{:>12,.0f}')} "
                  f"{_f(mape_m, '{:>9.1f}%')} "
                  f"{_fw(wins, n_comp):>9} "
                  f"{_f(win_pct, '{:>5.1f}%')} "
                  f"{_f(delta, '{:>+10.1f}%')} "
                  f"{_fp(p_wx):>10}")

    print(sep)
    print()


# Exportacao Excel

def _exportar_excel(
    df_full  : pd.DataFrame,
    resumo   : pd.DataFrame,
    path     : Path,
) -> None:
    """Gera Excel com 3 abas: Resumo, Completo, Vitorias."""
    # Colunas de RMSE para formatacao
    rmse_cols = [c for c in df_full.columns if "RMSE" in c]

    # Aba Vitorias: por serie, quem ganhou
    colunas_paradigma_rmse = {
        rot: col
        for rot, col in PARADIGMAS_RMSE.items()
        if col in df_full.columns and rot != "Individual"
    }

    vit_rows = []
    for _, row in df_full.iterrows():
        ind_rmse = row.get("Ind_RMSE", float("nan"))
        if isinstance(ind_rmse, float) and np.isnan(ind_rmse):
            continue
        vencedores = []
        for rot, col in colunas_paradigma_rmse.items():
            v = row.get(col, float("nan"))
            if not (isinstance(v, float) and np.isnan(v)):
                if v < ind_rmse:
                    vencedores.append(rot)
        vit_rows.append({
            "Codigo" : row.get("Codigo", ""),
            "Nome"   : row.get("Nome", ""),
            "Ind_RMSE": ind_rmse,
            "Melhor_Paradigma": min(
                colunas_paradigma_rmse.items(),
                key=lambda kv: row.get(kv[1], float("inf"))
                    if not np.isnan(row.get(kv[1], float("nan")))
                    else float("inf"),
                default=(None, None)
            )[0],
            "N_Paradigmas_Vencedores": len(vencedores),
            "Paradigmas_Vencedores"  : ", ".join(vencedores) if vencedores else "Individual",
        })

    vit_df = pd.DataFrame(vit_rows)

    # Exporta
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Aba 1: Resumo
        if not resumo.empty:
            resumo.to_excel(writer, sheet_name="Resumo", index=False)
            _formatar_aba_resumo(writer.sheets["Resumo"], resumo)

        # Aba 2: Completo
        df_full.to_excel(writer, sheet_name="Completo", index=False)
        ws_comp = writer.sheets["Completo"]
        _autofit_cols(ws_comp)
        for c_name in rmse_cols:
            if c_name in df_full.columns:
                ci = list(df_full.columns).index(c_name) + 1
                cl = _col_letter(ci)
                for r in range(2, len(df_full) + 2):
                    ws_comp[f"{cl}{r}"].number_format = "#,##0.00"

        # Aba 3: Vitorias
        if not vit_df.empty:
            vit_df.to_excel(writer, sheet_name="Vitorias", index=False)
            _autofit_cols(writer.sheets["Vitorias"])


def _formatar_aba_resumo(ws, resumo: pd.DataFrame) -> None:
    """Formata a aba Resumo com cores condicionais simples."""
    try:
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Destaca linha com menor RMSE_Medio (exceto Individual e Oracle)
        rmse_col_idx = None
        for i, col in enumerate(resumo.columns, 1):
            if col == "RMSE_Medio":
                rmse_col_idx = i
                break

        if rmse_col_idx:
            melhor_fill = PatternFill("solid", fgColor="C6EFCE")
            rmses = []
            for ri, row_data in resumo.iterrows():
                pard = str(row_data.get("Paradigma", ""))
                val  = row_data.get("RMSE_Medio", float("nan"))
                if pard not in ("Individual", "Oracle") and not np.isnan(val):
                    rmses.append((val, ri + 2))  # +2 offset de cabecalho
            if rmses:
                _, best_row = min(rmses, key=lambda x: x[0])
                for cell in ws[best_row]:
                    cell.fill = melhor_fill
    except Exception:
        pass

    _autofit_cols(ws)


def _col_letter(n: int) -> str:
    """Converte indice 1-based para letra de coluna Excel (A, B, ..., Z, AA, ...)."""
    s = ""
    while n:
        s = chr(64 + (n - 1) % 26 + 1) + s
        n = (n - 1) // 26
    return s


def _autofit_cols(ws) -> None:
    """Ajusta a largura das colunas automaticamente."""
    try:
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0)
                for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 40)
    except Exception:
        pass


# Entry point

if __name__ == "__main__":
    resultado = comparar_todos(verbose=True)
    if resultado is None:
        sys.exit(1)
